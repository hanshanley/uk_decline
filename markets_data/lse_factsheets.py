"""Fetch UK listed-company counts from London Stock Exchange monthly factsheets.

The World Bank / WFE "listed domestic companies" series (used elsewhere in ``markets_data``)
ends at 2022 for the UK. The LSE, however, publishes the underlying counts every month as
machine-readable factsheets on its documents host, e.g.:

    https://docs.londonstockexchange.com/sites/default/files/reports/Main%20Market%20factsheet%20December%202024.xlsx
    https://docs.londonstockexchange.com/sites/default/files/reports/AIM%20factsheet%20December%202024.xlsx

Each **Main Market** factsheet's "Summary" sheet gives that year's company count split UK /
International; each **AIM** factsheet's "AIM since launch" sheet gives the full annual history
(UK / International / Total). We read the December factsheet for each requested year and sum
Main Market + AIM to get total companies on the LSE, and (separately) the **UK-domiciled**
count -- which matches the World Bank/WFE "listed domestic companies" figure to within <1% at
the 2022 overlap (WB 1,606 vs LSE 1,619), so it is a near-seamless continuation of that series.

All figures are read straight from the LSE factsheets; nothing is hand-entered.
"""

from __future__ import annotations

import io
import re
from urllib.parse import urlsplit

import openpyxl
import requests

DOCS_HOST = "docs.londonstockexchange.com"
BASE = f"https://{DOCS_HOST}/sites/default/files/reports"
SOURCE = "London Stock Exchange monthly factsheets (Main Market + AIM), primary-market summary"
_UA = {"User-Agent": "uk_decline/0.1 (UK listed-company counts research)"}


def _ensure_lse_host(url: str) -> None:
    host = (urlsplit(url).hostname or "").lower()
    if not (host == DOCS_HOST or host.endswith(".londonstockexchange.com")):
        raise ValueError(f"refusing to fetch untrusted host: {url!r}")


def _factsheet_url(market: str, year: int, month: str = "December") -> str:
    # market is "Main Market" or "AIM"; spaces are URL-encoded by requests' params-free path.
    name = f"{market} factsheet {month} {year}.xlsx".replace(" ", "%20")
    return f"{BASE}/{name}"


def _download(url: str, timeout: int = 90) -> io.BytesIO:
    _ensure_lse_host(url)
    resp = requests.get(url, headers=_UA, timeout=timeout)
    _ensure_lse_host(resp.url)
    resp.raise_for_status()
    return io.BytesIO(resp.content)


def _num(x) -> int | None:
    try:
        return int(str(x).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def main_market_year(year: int, month: str = "December") -> tuple[int, int] | None:
    """Return ``(uk, international)`` Main Market company counts for ``year`` end, or None."""
    wb = openpyxl.load_workbook(_download(_factsheet_url("Main Market", year, month)),
                                data_only=True, read_only=True)
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(rows):
        vals = [str(c) for c in r if c is not None]
        if "Year" in vals and "Number of Companies" in vals:
            for r2 in rows[i:i + 4]:
                c = list(r2)
                if len(c) > 3 and _num(c[1]) == year:
                    return _num(c[2]), _num(c[3])
    return None


def aim_history(latest_year: int, month: str = "December") -> dict[int, tuple[int, int]]:
    """Return ``{year: (uk, international)}`` AIM counts from one factsheet's history sheet."""
    wb = openpyxl.load_workbook(_download(_factsheet_url("AIM", latest_year, month)),
                                data_only=True, read_only=True)
    ws = wb["AIM since launch"]
    out: dict[int, tuple[int, int]] = {}
    for r in ws.iter_rows(values_only=True):
        c = list(r)
        if len(c) > 4 and re.fullmatch(r"\d{4}", str(c[1]).strip() if c[1] is not None else ""):
            yr = int(c[1])
            uk, intl = _num(c[2]), _num(c[4])
            if uk is not None:
                out[yr] = (uk, intl or 0)
    return out


def build_rows(start: int = 2022, end: int = 2025) -> list[dict]:
    """Tidy rows of LSE company counts per year: total and UK-domiciled (Main Market + AIM)."""
    aim = aim_history(end)
    rows: list[dict] = []
    for year in range(start, end + 1):
        mm = main_market_year(year)
        if mm is None or year not in aim:
            continue
        mm_uk, mm_intl = mm[0] or 0, mm[1] or 0
        aim_uk, aim_intl = aim[year]
        uk = mm_uk + aim_uk
        total = mm_uk + mm_intl + aim_uk + aim_intl
        src_url = _factsheet_url("Main Market", year).replace("%20", " ")
        for metric, value in (("lse_companies_uk", uk), ("lse_companies_total", total)):
            rows.append({
                "region": "United Kingdom",
                "year": year,
                "metric": metric,
                "value": value,
                "unit": "companies",
                "source": SOURCE,
                "source_url": src_url,
            })
    rows.sort(key=lambda r: (r["metric"], r["year"]))
    return rows


def write_csv(path: str | None = None, start: int = 2022, end: int = 2025) -> str:
    """Fetch the LSE factsheets and write the tidy cited CSV. Returns the path."""
    import csv
    import pathlib

    path = path or str(pathlib.Path(__file__).resolve().parent.parent
                       / "data" / "uk_listed_companies_lse.csv")
    rows = build_rows(start, end)
    fields = ["region", "year", "metric", "value", "unit", "source", "source_url"]
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    return path


if __name__ == "__main__":
    import datetime

    out = write_csv(end=datetime.date.today().year - 1)
    print(f"[lse_factsheets] wrote {out}")
