"""Read tabular rows from a downloaded spreadsheet (xlsx / xls / ods / csv).

Home Office and ONS publish machine-readable ``Data_*`` / numbered sheets. These helpers
stream a named sheet into row tuples so the source parsers can aggregate without loading
everything into memory (some datasets have hundreds of thousands of rows). Legacy ONS
adhoc tables are old-format ``.xls`` (read via ``xlrd``).
"""

from __future__ import annotations

import csv
import io
from typing import Iterator


def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def read_rows(content: bytes, filename: str, sheet: str | None = None) -> Iterator[tuple]:
    """Yield row tuples from ``content``, dispatching on the ``filename`` extension.

    ``sheet`` selects a worksheet for xlsx/ods; it is ignored for csv.
    """
    ext = _ext(filename)
    if ext == "xlsx":
        yield from _read_xlsx(content, sheet)
    elif ext == "xls":
        yield from _read_xls(content, sheet)
    elif ext == "ods":
        yield from _read_ods(content, sheet)
    elif ext == "csv":
        yield from _read_csv(content)
    else:
        raise ValueError(f"unsupported spreadsheet type {ext!r} for {filename}")


def _read_xls(content: bytes, sheet: str | None) -> Iterator[tuple]:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_name(sheet) if sheet else wb.sheet_by_index(0)
    for r in range(ws.nrows):
        yield tuple(ws.cell_value(r, c) for c in range(ws.ncols))


def _read_xlsx(content: bytes, sheet: str | None) -> Iterator[tuple]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield row
    finally:
        wb.close()


def _read_ods(content: bytes, sheet: str | None) -> Iterator[tuple]:
    import pandas as pd

    frame = pd.read_excel(
        io.BytesIO(content), sheet_name=sheet or 0, engine="odf", header=None
    )
    for row in frame.itertuples(index=False, name=None):
        yield row


def _read_csv(content: bytes) -> Iterator[tuple]:
    text = content.decode("utf-8-sig", errors="replace")
    for row in csv.reader(io.StringIO(text)):
        yield tuple(row)
